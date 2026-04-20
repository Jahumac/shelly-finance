from datetime import date, datetime
from io import BytesIO

from flask import Blueprint, flash, jsonify, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required

from app.utils import optional_float, optional_int, valid_month_key

from app.models import (
    build_debt_card,
    create_budget_item,
    create_budget_section,
    create_debt,
    delete_budget_item,
    delete_budget_items_by_section,
    delete_budget_section,
    delete_debt,
    fetch_all_accounts,
    fetch_all_debts,
    fetch_budget_entries,
    fetch_budget_item,
    fetch_budget_items,
    fetch_budget_sections,
    fetch_budget_trend,
    fetch_debt,
    fetch_months_with_budget_entries,
    fetch_prior_month_budget_entries,
    update_budget_item,
    update_budget_section,
    update_debt,
    upsert_budget_entry,
)
from app.models.debts import amortisation_schedule

budget_bp = Blueprint("budget", __name__)


def _default_month_key():
    today = date.today()
    return f"{today.year}-{today.month:02d}"


def _month_label(month_key):
    return datetime.strptime(month_key, "%Y-%m").strftime("%B %Y")


def _build_monthly_data(month_key, user_id):
    db_sections = fetch_budget_sections(user_id)
    items = fetch_budget_items(user_id)
    entries = fetch_budget_entries(month_key, user_id)
    entry_map = {e["budget_item_id"]: e for e in entries}
    accounts = fetch_all_accounts(user_id)
    account_map = {a["id"]: a for a in accounts}

    # Always load prior-month entries so we can show per-item inheritance
    prior_entries = fetch_prior_month_budget_entries(month_key, user_id)
    prior_entry_map = {e["budget_item_id"]: e for e in prior_entries}

    _income_sec = next((s for s in db_sections if "income" in s["key"].lower()), None)
    income_key = _income_sec["key"] if _income_sec else (db_sections[0]["key"] if db_sections else "income")

    sections = []
    section_totals = {}

    for sec in db_sections:
        section_key = sec["key"]
        section_items = []
        for item in items:
            if item["section"] != section_key:
                continue
            if item["id"] in entry_map:
                amount = float(entry_map[item["id"]]["amount"] or 0)
            elif item["id"] in prior_entry_map:
                amount = float(prior_entry_map[item["id"]]["amount"] or 0)
            elif item["linked_account_id"] and item["linked_account_id"] in account_map:
                amount = float(account_map[item["linked_account_id"]]["monthly_contribution"] or 0)
            else:
                amount = float(item["default_amount"] or 0)
            # Track whether amount came from the current month or was inherited
            source = "current"
            if item["id"] in entry_map:
                source = "current"
            elif item["id"] in prior_entry_map:
                source = "inherited"
            elif item["linked_account_id"] and item["linked_account_id"] in account_map:
                source = "linked"
            else:
                source = "default"

            linked_account_name = None
            if item["linked_account_id"] and item["linked_account_id"] in account_map:
                linked_account_name = account_map[item["linked_account_id"]]["name"]

            section_items.append({
                "id": item["id"],
                "name": item["name"],
                "notes": item["notes"],
                "amount": amount,
                "linked": item["linked_account_id"] is not None,
                "linked_account_name": linked_account_name,
                "source": source,
            })
        section_total = sum(i["amount"] for i in section_items)
        section_totals[section_key] = section_total
        sections.append({
            "key": section_key,
            "label": sec["label"],
            "rows": section_items,
            "total": section_total,
        })

    total_income = section_totals.get(income_key, 0)
    total_expenses = sum(v for k, v in section_totals.items() if k != income_key)
    surplus = total_income - total_expenses
    savings_total = 0.0
    for sec in db_sections:
        k = sec["key"]
        if k != income_key and ("invest" in k or "saving" in k):
            savings_total += section_totals.get(k, 0)
    savings_rate = (savings_total / total_income * 100) if total_income > 0 else 0.0

    summary = {
        "total_income": total_income,
        "total_expenses": total_expenses,
        "surplus": surplus,
        "savings_rate": savings_rate,
    }

    return sections, summary


@budget_bp.route("/", methods=["GET", "POST"])
@login_required
def budget():
    uid = current_user.id
    month_key = valid_month_key(request.values.get("month")) or _default_month_key()

    if request.method == "POST":
        item_id = request.form.get("item_id", type=int)
        if item_id:
            upsert_budget_entry(month_key, item_id, optional_float(request.form.get("amount"), 0.0), uid)
        return redirect(url_for("budget.budget", month=month_key))

    sections, summary = _build_monthly_data(month_key, uid)
    db_sections = fetch_budget_sections(uid)
    _income_sec = next((s for s in db_sections if "income" in s["key"].lower()), None)
    income_key = _income_sec["key"] if _income_sec else (db_sections[0]["key"] if db_sections else "income")

    month_strip, is_inherited = _build_budget_month_strip(month_key, uid)

    return render_template(
        "budget.html",
        month_key=month_key,
        month_label=_month_label(month_key),
        sections=sections,
        summary=summary,
        income_key=income_key,
        month_strip=month_strip,
        is_inherited=is_inherited,
        active_page="budget",
    )


@budget_bp.route("/api/entry", methods=["POST"])
@login_required
def budget_save_entry():
    """AJAX endpoint — save a single budget entry, return JSON."""
    uid = current_user.id
    month_key = valid_month_key(request.form.get("month")) or _default_month_key()
    item_id = request.form.get("item_id", type=int)
    amount = optional_float(request.form.get("amount"), 0.0)
    if item_id:
        upsert_budget_entry(month_key, item_id, amount, uid)
    return jsonify({"ok": True})


@budget_bp.route("/api/quick-add", methods=["POST"])
@login_required
def budget_quick_add():
    """AJAX endpoint — add a new budget item from the budget view."""
    uid = current_user.id
    name = (request.form.get("name") or "").strip()
    section = (request.form.get("section") or "").strip()
    if not name or not section:
        return jsonify({"ok": False, "error": "Name and section required"}), 400

    existing = fetch_budget_items(uid)
    sort_order = max(
        (i["sort_order"] for i in existing if i["section"] == section), default=-1
    ) + 1
    item_id = create_budget_item({
        "name": name,
        "section": section,
        "default_amount": 0.0,
        "linked_account_id": None,
        "notes": "",
        "sort_order": sort_order,
    }, uid)
    return jsonify({"ok": True, "item_id": item_id, "name": name})


@budget_bp.route("/import", methods=["POST"])
@login_required
def budget_import():
    """Import budget items and amounts from an uploaded .xlsx file.

    Reads the Shelly export format:
      Row 1: title
      Row 2: generated date
      Then repeating blocks of:
        - Section header row (col A = section label, col C = "Amount")
        - Item rows (col A = name, col B = notes, col C = amount)
        - "Section total" row
        - Blank rows
      Finally summary rows (Total Income, Total Expenses, Surplus)
    """
    uid = current_user.id
    month_key = valid_month_key(request.form.get("month")) or _default_month_key()

    f = request.files.get("file")
    if not f or not f.filename.endswith((".xlsx", ".xls")):
        flash("Please upload an .xlsx file.", "error")
        return redirect(url_for("budget.budget", month=month_key))

    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception:
        flash("Could not read the Excel file.", "error")
        return redirect(url_for("budget.budget", month=month_key))

    # Load existing sections and items for matching
    db_sections = fetch_budget_sections(uid)
    section_label_to_key = {s["label"].strip().lower(): s["key"] for s in db_sections}
    existing_items = fetch_budget_items(uid)

    current_section_key = None
    items_imported = 0
    sections_created = 0
    skip_labels = {"total income", "total expenses", "surplus", "section total", ""}

    for row in ws.iter_rows(min_row=1, values_only=False):
        vals = [cell.value for cell in row]
        col_a = str(vals[0] or "").strip()
        col_b = str(vals[1] or "").strip()
        col_c = vals[2] if len(vals) > 2 else None

        # Skip title, date, and empty rows
        if not col_a and not col_c:
            continue

        # Detect section header: col C is literally "Amount"
        if isinstance(col_c, str) and col_c.strip().lower() == "amount":
            label = col_a
            label_lower = label.lower()
            if label_lower in section_label_to_key:
                current_section_key = section_label_to_key[label_lower]
            else:
                # Create a new section
                new_key = create_budget_section(label, uid)
                section_label_to_key[label_lower] = new_key
                current_section_key = new_key
                sections_created += 1
                # Refresh sections
                db_sections = fetch_budget_sections(uid)
            continue

        # Skip summary/total rows
        if col_a.lower() in skip_labels or col_b.lower() == "section total":
            continue

        # Skip rows without a numeric amount
        amount = None
        if isinstance(col_c, (int, float)):
            amount = float(col_c)
        else:
            try:
                cleaned = str(col_c or "").replace("£", "").replace(",", "").strip()
                amount = float(cleaned) if cleaned else None
            except (ValueError, TypeError):
                amount = None

        if amount is None or not current_section_key or not col_a:
            continue

        # Find or create the budget item
        item_name = col_a
        notes = col_b if col_b else ""
        matched_item = None
        for it in existing_items:
            if it["name"].strip().lower() == item_name.lower() and it["section"] == current_section_key:
                matched_item = it
                break

        if not matched_item:
            # Create a new item
            sort_order = max(
                (i["sort_order"] for i in existing_items if i["section"] == current_section_key),
                default=-1,
            ) + 1
            item_id = create_budget_item({
                "name": item_name,
                "section": current_section_key,
                "default_amount": amount,
                "linked_account_id": None,
                "notes": notes,
                "sort_order": sort_order,
            }, uid)
            # Refresh items list
            existing_items = fetch_budget_items(uid)
        else:
            item_id = matched_item["id"]

        # Upsert the entry for this month
        upsert_budget_entry(month_key, item_id, amount, uid)
        items_imported += 1

    msg = f"Imported {items_imported} budget items"
    if sections_created:
        msg += f" and created {sections_created} new section{'s' if sections_created > 1 else ''}"
    msg += f" for {_month_label(month_key)}."
    flash(msg, "success")
    return redirect(url_for("budget.budget", month=month_key))


@budget_bp.route("/trend/")
@login_required
def budget_trend():
    uid = current_user.id
    today = date.today()

    # Last 6 months that have any entries
    all_months = _last_n_months(today, 6)
    months_with_data = fetch_months_with_budget_entries(uid)
    months = [m for m in all_months if m in months_with_data]

    if not months:
        return render_template(
            "budget_trend.html",
            sections={},
            months=[],
            month_labels=[],
            current_month_num=today.month,
            trend_avg_income=0,
            trend_avg_spend=0,
            trend_surplus=0,
            active_page="budget",
        )

    # Load all active items, sections, accounts (for linked defaults)
    db_sections = fetch_budget_sections(uid)
    all_items = fetch_budget_items(uid)
    accounts = fetch_all_accounts(uid)
    account_map = {a["id"]: a for a in accounts}
    section_label = {s["key"]: s["label"] for s in db_sections}
    section_order = {s["key"]: s["sort_order"] for s in db_sections}

    # Build entry lookup: {month_key: {item_id: amount}}
    entry_lookup = {}
    for mk in months:
        entry_lookup[mk] = {
            e["budget_item_id"]: float(e["amount"] or 0)
            for e in fetch_budget_entries(mk, uid)
        }

    # Build structure using ALL active items, falling back to linked/default amounts
    sections = {}
    for item in all_items:
        skey = item["section"]
        sn = section_label.get(skey, skey)
        inn = item["name"]

        # Default amount: prefer linked account monthly contribution
        if item["linked_account_id"] and item["linked_account_id"] in account_map:
            fallback = float(account_map[item["linked_account_id"]]["monthly_contribution"] or 0)
        else:
            fallback = float(item["default_amount"] or 0)

        if sn not in sections:
            sections[sn] = {"_order": section_order.get(skey, 99)}
        if inn not in sections[sn]:
            sections[sn][inn] = {"default": fallback}

        for mk in months:
            if item["id"] in entry_lookup.get(mk, {}):
                sections[sn][inn][mk] = entry_lookup[mk][item["id"]]
            else:
                sections[sn][inn][mk] = fallback

    # Sort sections by sort_order, remove internal _order key
    sections = {
        k: {ik: iv for ik, iv in v.items() if ik != "_order"}
        for k, v in sorted(sections.items(), key=lambda x: x[1].get("_order", 99))
    }

    # Averages
    for sn, items in sections.items():
        for inn, data in items.items():
            month_vals = [data[mk] for mk in months if mk in data]
            data["avg"] = sum(month_vals) / len(month_vals) if month_vals else 0

    month_labels = [
        datetime.strptime(mk, "%Y-%m").strftime("%b %Y") for mk in months
    ]

    # Hero stats
    income_section = next((s for s in db_sections if "income" in s["key"].lower()), None)
    income_key = income_section["key"] if income_section else None
    income_section_label = income_section["label"] if income_section else None

    trend_avg_income = 0.0
    trend_avg_spend = 0.0
    trend_item_count = 0
    for sn, items in sections.items():
        is_income = (income_section_label and sn == income_section_label)
        for inn, data in items.items():
            trend_item_count += 1
            if is_income:
                trend_avg_income += data["avg"]
            else:
                trend_avg_spend += data["avg"]

    return render_template(
        "budget_trend.html",
        sections=sections,
        months=months,
        month_labels=month_labels,
        trend_avg_income=trend_avg_income,
        trend_avg_spend=trend_avg_spend,
        trend_surplus=trend_avg_income - trend_avg_spend,
        trend_item_count=trend_item_count,
        active_page="budget",
    )


def _build_budget_month_strip(month_key, uid):
    """Build the 12-month tax-year strip (Apr→Mar) for the budget views."""
    saved_months = fetch_months_with_budget_entries(uid)
    mk_year, mk_month = int(month_key[:4]), int(month_key[5:7])
    ty_start_year = mk_year if mk_month >= 4 else mk_year - 1
    today_key = _default_month_key()

    month_strip = []
    for i in range(12):
        m = 4 + i
        y = ty_start_year if m <= 12 else ty_start_year + 1
        if m > 12:
            m -= 12
        mk = f"{y}-{m:02d}"
        month_strip.append({
            "key": mk,
            "label": datetime.strptime(mk, "%Y-%m").strftime("%b"),
            "has_data": mk in saved_months,
            "is_current": mk == month_key,
            "is_today": mk == today_key,
            "month_num": m,
        })

    return month_strip, month_key not in saved_months  # (strip, is_inherited)


def _last_n_months(today, n):
    """Return list of 'YYYY-MM' strings for the last n months (most recent last)."""
    result = []
    y, m = today.year, today.month
    for _ in range(n):
        result.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    result.reverse()
    return result


@budget_bp.route("/items/", methods=["GET", "POST"])
@login_required
def budget_items_view():
    uid = current_user.id
    if request.method == "POST":
        form_name = request.form.get("form_name", "")

        if form_name == "clear_section":
            delete_budget_items_by_section(request.form.get("section_key", ""), uid)
            return redirect(url_for("budget.budget_items_view"))

        if form_name == "add_section":
            label = request.form.get("section_label", "").strip()
            if label:
                create_budget_section(label, uid)
            return redirect(url_for("budget.budget_items_view"))

        if form_name == "edit_section":
            update_budget_section(
                request.form.get("section_key", ""),
                request.form.get("section_label", "").strip(),
                uid,
            )
            return redirect(url_for("budget.budget_items_view"))

        if form_name == "delete_section":
            delete_budget_section(request.form.get("section_key", ""), uid)
            return redirect(url_for("budget.budget_items_view"))

        # default: create item
        section = request.form.get("section", "")
        existing = fetch_budget_items(uid)
        sort_order = max(
            (i["sort_order"] for i in existing if i["section"] == section), default=-1
        ) + 1
        linked_raw = request.form.get("linked_account_id", "")
        create_budget_item({
            "name": request.form.get("name", "").strip(),
            "section": section,
            "default_amount": optional_float(request.form.get("default_amount"), 0.0),
            "linked_account_id": int(linked_raw) if linked_raw else None,
            "notes": request.form.get("notes", "").strip(),
            "sort_order": sort_order,
        }, uid)
        return redirect(url_for("budget.budget_items_view"))

    db_sections = fetch_budget_sections(uid)
    all_items = fetch_budget_items(uid)
    accounts = fetch_all_accounts(uid)
    account_map = {a["id"]: dict(a) for a in accounts}

    grouped = []
    for sec in db_sections:
        section_key = sec["key"]
        section_items = []
        for item in all_items:
            if item["section"] != section_key:
                continue
            row = dict(item)
            row["linked_account_name"] = (
                account_map[item["linked_account_id"]]["name"]
                if item["linked_account_id"] and item["linked_account_id"] in account_map
                else None
            )
            section_items.append(row)
        grouped.append({"key": section_key, "label": sec["label"], "rows": section_items})

    selected_id = request.args.get("item_id", type=int)
    selected = fetch_budget_item(selected_id, uid) if selected_id else None
    page_mode = request.args.get("mode", "view" if selected_id else "list")
    section_options = [(s["key"], s["label"]) for s in db_sections]

    # ── Month context (same hero + month strip as budget view) ────────────
    month_key = valid_month_key(request.args.get("month")) or _default_month_key()
    sections_data, summary = _build_monthly_data(month_key, uid)
    month_strip, is_inherited = _build_budget_month_strip(month_key, uid)

    return render_template(
        "budget_items.html",
        grouped=grouped,
        accounts=accounts,
        section_options=section_options,
        selected=selected,
        page_mode=page_mode,
        active_page="budget",
        month_key=month_key,
        month_label=_month_label(month_key),
        summary=summary,
        month_strip=month_strip,
        is_inherited=is_inherited,
    )


@budget_bp.route("/items/<int:item_id>", methods=["POST"])
@login_required
def budget_item_action(item_id):
    uid = current_user.id
    if request.form.get("form_name") == "delete":
        delete_budget_item(item_id, uid)
        return redirect(url_for("budget.budget_items_view"))

    linked_raw = request.form.get("linked_account_id", "")
    ok = update_budget_item({
        "id": item_id,
        "name": request.form.get("name", "").strip(),
        "section": request.form.get("section", ""),
        "default_amount": max(0.0, optional_float(request.form.get("default_amount"), 0.0)),
        "linked_account_id": int(linked_raw) if linked_raw else None,
        "notes": request.form.get("notes", "").strip(),
    }, uid)
    if not ok:
        flash("Budget item not found.", "error")
    return redirect(url_for("budget.budget_items_view"))


# ── Debts ─────────────────────────────────────────────────────────────────────

@budget_bp.route("/debts/", methods=["GET", "POST"])
@login_required
def budget_debts():
    uid = current_user.id

    if request.method == "POST":
        form = request.form
        form_name = form.get("form_name")

        if form_name == "create_debt":
            create_debt({
                "name": form.get("name", "").strip(),
                "original_amount": optional_float(form.get("original_amount"), 0.0),
                "current_balance": optional_float(form.get("current_balance"), 0.0),
                "monthly_payment": optional_float(form.get("monthly_payment"), 0.0),
                "apr": optional_float(form.get("apr"), 0.0),
                "notes": form.get("notes", "").strip(),
                "start_date": form.get("start_date", "").strip() or None,
            }, uid)
            return redirect(url_for("budget.budget_debts"))

        if form_name == "update_debt":
            debt_id = optional_int(form.get("debt_id"))
            if debt_id and fetch_debt(debt_id, uid):
                update_debt(debt_id, {
                    "name": form.get("name", "").strip(),
                    "original_amount": optional_float(form.get("original_amount"), 0.0),
                    "current_balance": optional_float(form.get("current_balance"), 0.0),
                    "monthly_payment": optional_float(form.get("monthly_payment"), 0.0),
                    "apr": optional_float(form.get("apr"), 0.0),
                    "notes": form.get("notes", "").strip(),
                    "start_date": form.get("start_date", "").strip() or None,
                }, uid)
            return redirect(url_for("budget.budget_debts"))

        if form_name == "delete_debt":
            debt_id = optional_int(form.get("debt_id"))
            if debt_id and fetch_debt(debt_id, uid):
                delete_debt(debt_id, uid)
            return redirect(url_for("budget.budget_debts"))

        return redirect(url_for("budget.budget_debts"))

    raw_debts = fetch_all_debts(uid)
    debt_cards = [build_debt_card(d) for d in raw_debts]

    selected_id = request.args.get("debt_id", type=int)
    page_mode = request.args.get("mode", "view")
    selected_debt_raw = next((d for d in raw_debts if d["id"] == selected_id), None) if selected_id else None
    selected_debt = next((d for d in debt_cards if d["id"] == selected_id), None) if selected_id else None

    # Build amortisation schedule for the selected debt detail view.
    # If auto-tracked (start_date + original_amount set), show the full schedule
    # from loan inception so past payments are visible alongside future ones.
    schedule = []
    payments_made = 0
    total_interest_all = None
    interest_paid = 0

    if selected_debt:
        payments_made = selected_debt.get("payments_made", 0)
        original = selected_debt.get("original_amount", 0)

        if selected_debt["auto_tracked"] and original > 0:
            # Full schedule from original balance
            schedule = amortisation_schedule(
                original,
                selected_debt["apr"],
                selected_debt["monthly_payment"],
            )
        elif selected_debt["months_remaining"]:
            # Remaining schedule only (no start date)
            schedule = amortisation_schedule(
                selected_debt["current_balance"],
                selected_debt["apr"],
                selected_debt["monthly_payment"],
            )

        if schedule:
            total_interest_all = sum(r["interest"] for r in schedule)
            interest_paid = sum(r["interest"] for r in schedule[:payments_made])
        else:
            interest_paid = 0

    return render_template(
        "budget_debts.html",
        debt_cards=debt_cards,
        selected_debt=selected_debt,
        selected_debt_raw=selected_debt_raw,
        schedule=schedule,
        payments_made=payments_made,
        total_interest_all=total_interest_all,
        interest_paid=interest_paid,
        page_mode=page_mode,
        active_page="budget",
    )


@budget_bp.route("/debts/export.xlsx")
@login_required
def budget_debts_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Shelly style constants (match projections export) ────────────────────
    TEAL       = "0F766E"
    TEAL_LIGHT = "CCFBF1"
    RED_LIGHT  = "FEE2E2"
    GREEN_LIGHT= "DCFCE7"
    BORDER_CLR = "D1D5DB"

    title_font  = Font(name="Aptos", bold=True, color=TEAL, size=14)
    sub_font    = Font(name="Aptos", color="6B7280", size=10)
    hdr_font    = Font(name="Aptos", bold=True, color="FFFFFF", size=11)
    data_font   = Font(name="Aptos", color="1F2937", size=10)
    bold_font   = Font(name="Aptos", bold=True, color="1F2937", size=10)
    red_font    = Font(name="Aptos", color="DC2626", size=10)
    green_font  = Font(name="Aptos", color="16A34A", size=10)

    hdr_fill    = PatternFill("solid", fgColor=TEAL)
    alt_fill    = PatternFill("solid", fgColor=TEAL_LIGHT)
    red_fill    = PatternFill("solid", fgColor=RED_LIGHT)
    green_fill  = PatternFill("solid", fgColor=GREEN_LIGHT)
    no_fill     = PatternFill(fill_type=None)
    thin_border = Border(bottom=Side(style="thin", color=BORDER_CLR))
    GBP         = '£#,##0.00'
    GBP0        = '£#,##0'

    def col_width(ws, col, w):
        ws.column_dimensions[get_column_letter(col)].width = w

    def hdr_row(ws, row, values, widths=None):
        for i, v in enumerate(values, 1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(vertical="center", horizontal="left")
        ws.row_dimensions[row].height = 22
        if widths:
            for i, w in enumerate(widths, 1):
                col_width(ws, i, w)

    def data_cell(ws, row, col, value, font=None, fill=None, num_fmt=None, bold=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font or (bold_font if bold else data_font)
        c.fill = fill or (alt_fill if row % 2 == 0 else no_fill)
        c.border = thin_border
        c.alignment = Alignment(vertical="center")
        if num_fmt:
            c.number_format = num_fmt
        return c

    uid = current_user.id
    raw_debts = fetch_all_debts(uid)
    debt_cards = [build_debt_card(d) for d in raw_debts]

    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 1 — Summary
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"

    c = ws.cell(row=1, column=1, value="Shelly Finance — Debt Tracker")
    c.font = title_font
    ws.merge_cells("A1:H1")

    c = ws.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%d %b %Y at %H:%M')}")
    c.font = sub_font
    ws.merge_cells("A2:H2")
    ws.row_dimensions[1].height = 28

    hdr_row(ws, 4,
            ["Debt", "Balance", "Monthly Payment", "APR %",
             "Months Left", "Payoff Date", "Total Interest", "Total Cost"],
            widths=[28, 16, 18, 10, 14, 16, 18, 16])

    for i, d in enumerate(debt_cards, 5):
        fill = alt_fill if i % 2 == 0 else no_fill
        data_cell(ws, i, 1, d["name"],        bold=True)
        data_cell(ws, i, 2, d["current_balance"],   num_fmt=GBP)
        data_cell(ws, i, 3, d["monthly_payment"],    num_fmt=GBP)
        data_cell(ws, i, 4, d["apr"],                num_fmt='0.00"%"')
        data_cell(ws, i, 5, d["months_remaining"])
        data_cell(ws, i, 6, d["payoff_date"].strftime("%b %Y") if d["payoff_date"] else "—")
        c = data_cell(ws, i, 7, d["total_interest"],     num_fmt=GBP0)
        c.font = red_font
        data_cell(ws, i, 8, d["total_cost"],         num_fmt=GBP0)

    ws.freeze_panes = "A5"

    # ════════════════════════════════════════════════════════════════════════
    # Per-debt sheets — one tab per scenario
    # ════════════════════════════════════════════════════════════════════════
    for d in debt_cards:
        if not d["months_remaining"]:
            continue

        base_payment = d["monthly_payment"]
        scenarios = [
            ("Base",    0),
            ("+£50",   50),
            ("+£100", 100),
            ("+£200", 200),
            ("Double", base_payment),
        ]

        for label, extra in scenarios:
            new_payment = base_payment + extra
            sched = amortisation_schedule(d["current_balance"], d["apr"], new_payment)
            total_interest = sum(r["interest"] for r in sched)
            months_saved = (d["months_remaining"] or 0) - len(sched) if extra > 0 else 0
            interest_saved = (d["total_interest"] or 0) - total_interest if extra > 0 else 0

            # Tab name: e.g. "Car Loan — Base", "Car Loan — +£50"
            tab = f"{d['name'][:20]} — {label}"
            ws2 = wb.create_sheet(title=tab)

            # Title
            title_str = f"{d['name']} — {label}"
            if extra > 0:
                title_str += f" (£{new_payment:,.2f}/mo)"
            c = ws2.cell(row=1, column=1, value=f"Shelly Finance — {title_str}")
            c.font = title_font
            ws2.merge_cells("A1:E1")
            ws2.row_dimensions[1].height = 28

            # Summary box rows 2-5
            summaries = [
                ("Balance",        f"£{d['current_balance']:,.2f}"),
                ("Monthly payment",f"£{new_payment:,.2f}"),
                ("Payoff in",      f"{len(sched)} months"),
                ("Total interest", f"£{total_interest:,.0f}"),
            ]
            if extra > 0:
                summaries += [
                    ("Months saved",   f"{months_saved}"),
                    ("Interest saved", f"£{interest_saved:,.0f}"),
                ]
            for r_i, (k, v) in enumerate(summaries, 2):
                lc = ws2.cell(row=r_i, column=1, value=k)
                lc.font = sub_font
                vc = ws2.cell(row=r_i, column=2, value=v)
                vc.font = bold_font if "saved" not in k else Font(name="Aptos", bold=True, color="16A34A", size=10)

            header_row = len(summaries) + 3
            hdr_row(ws2, header_row,
                    ["Month", "Payment", "Interest", "To Principal", "Balance"],
                    widths=[10, 16, 16, 16, 16])

            for row_i, row in enumerate(sched, header_row + 1):
                fill = alt_fill if row_i % 2 == 0 else no_fill
                data_cell(ws2, row_i, 1, row["month"])
                data_cell(ws2, row_i, 2, row["payment"],   num_fmt=GBP)
                c = data_cell(ws2, row_i, 3, row["interest"],  num_fmt=GBP)
                c.font = red_font
                c = data_cell(ws2, row_i, 4, row["principal"], num_fmt=GBP)
                c.font = green_font
                data_cell(ws2, row_i, 5, row["balance"],   num_fmt=GBP, bold=True)

            ws2.freeze_panes = f"A{header_row + 1}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="debts.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
