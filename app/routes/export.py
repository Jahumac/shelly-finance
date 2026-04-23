"""
Export routes — generates .xlsx downloads for Projections and Budget.

Clean, professional styling with Shelly-themed headers and plain UK-formatted data.
"""
from datetime import date, datetime
from io import BytesIO

from flask import Blueprint, send_file, request
from app.utils import valid_month_key
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Alignment, Side, numbers
from openpyxl.utils import get_column_letter

from app.calculations import (
    _safe_get,
    account_gross_growth_rate,
    account_growth_rate,
    compute_performance_series,
    contribution_breakdown,
    current_age_from_assumptions,
    effective_account_value,
    effective_fee_pct,
    effective_monthly_contribution,
    future_value,
    is_pension_account,
    projected_account_value,
    projected_account_value_at_year,
    projected_account_value_at_month,
    projected_account_value_at_month_no_fees,
    projected_account_value_at_year_no_fees,
    projected_account_value_no_fees,
    projected_total_retirement_value,
    to_float,
    uk_tax_year_end,
    uk_tax_year_start,
    years_to_retirement,
    ISA_WRAPPER_TYPES,
    LISA_WRAPPER_TYPES,
)
from app.models import (
    fetch_all_accounts,
    fetch_assumptions,
    fetch_budget_entries,
    fetch_budget_items,
    fetch_budget_sections,
    fetch_holding_totals_by_account,
    fetch_isa_contributions,
    fetch_monthly_performance_data,
    fetch_monthly_performance_data_by_account,
    fetch_pension_contributions,
    fetch_prior_month_budget_entries,
)

export_bp = Blueprint("export", __name__)

# ── Clean Shelly colour palette ──────────────────────────────────────────────
_SHELLY_TEAL   = "0F766E"   # Shelly's signature teal (header bg)
_SHELLY_LIGHT  = "CCFBF1"   # Pale teal tint for alternating rows
_BORDER_COLOUR = "D1D5DB"   # Light grey border

_TITLE_FONT    = Font(name="Aptos", bold=True, color="0F766E", size=14)
_SUBTITLE_FONT = Font(name="Aptos", color="6B7280", size=10)
_HEADER_FONT   = Font(name="Aptos", bold=True, color="FFFFFF", size=11)
_DATA_FONT     = Font(name="Aptos", color="1F2937", size=10)
_DATA_BOLD     = Font(name="Aptos", bold=True, color="1F2937", size=10)
_ACCENT_FONT   = Font(name="Aptos", bold=True, color="0F766E", size=11)

_HEADER_FILL   = PatternFill("solid", fgColor=_SHELLY_TEAL)
_ALT_FILL      = PatternFill("solid", fgColor=_SHELLY_LIGHT)
_NO_FILL       = PatternFill(fill_type=None)

_THIN_BORDER   = Border(
    bottom=Side(style="thin", color=_BORDER_COLOUR),
)


def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def _header_row(ws, row_num, values):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[row_num].height = 24


def _data_row(ws, row_num, values, bold=False, num_formats=None):
    num_formats = num_formats or {}
    font = _DATA_BOLD if bold else _DATA_FONT
    fill = _ALT_FILL if row_num % 2 == 0 else _NO_FILL
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
        cell.border = _THIN_BORDER
        if col in num_formats:
            cell.number_format = num_formats[col]


def _title_cell(ws, row_num, text, col_span=1):
    cell = ws.cell(row=row_num, column=1, value=text)
    cell.font = _TITLE_FONT
    if col_span > 1:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_span)


# ── Projections export ────────────────────────────────────────────────────────

@export_bp.route("/projections/export.xlsx")
@login_required
def export_projections():
    uid = current_user.id
    raw_accounts = fetch_all_accounts(uid)
    assumptions  = fetch_assumptions(uid)
    holdings_totals = fetch_holding_totals_by_account(uid)

    accounts = []
    for a in raw_accounts:
        row = dict(a)
        row["current_value"] = effective_account_value(a, holdings_totals)
        accounts.append(row)

    current_age    = current_age_from_assumptions(assumptions) if assumptions else 43
    retirement_age = to_float(assumptions["retirement_age"]) if assumptions else 60
    growth_rate    = to_float(assumptions["annual_growth_rate"]) if assumptions else 0.07
    exact_years    = years_to_retirement(current_age, retirement_age, assumptions) if assumptions else max(retirement_age - current_age, 0)
    whole_years    = int(exact_years)
    total_projected = projected_total_retirement_value(accounts, assumptions)

    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    _set_col_width(ws, 1, 34)
    _set_col_width(ws, 2, 18)
    _set_col_width(ws, 3, 18)
    _set_col_width(ws, 4, 24)

    _title_cell(ws, 1, "Shelly Finance — Retirement Projections", 4)
    cell = ws.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%d %b %Y at %H:%M')}")
    cell.font = _SUBTITLE_FONT

    _header_row(ws, 4, ["Account", "Current Value", "Monthly into Pot", "Projected at Retirement"])

    # UK pound formats
    GBP  = '£#,##0.00'
    GBP0 = '£#,##0'

    for i, acc in enumerate(accounts, 5):
        proj = projected_account_value(acc, assumptions)
        effective = effective_monthly_contribution(acc, assumptions)
        _data_row(ws, i, [
            acc["name"],
            to_float(acc["current_value"]),
            effective,
            proj,
        ], num_formats={2: GBP, 3: GBP, 4: GBP0})

    total_row = len(accounts) + 5
    _data_row(ws, total_row, [
        "Total",
        sum(to_float(a["current_value"]) for a in accounts),
        sum(effective_monthly_contribution(a, assumptions) for a in accounts),
        total_projected,
    ], bold=True, num_formats={2: GBP, 3: GBP, 4: GBP0})

    # Fee impact summary (only if any account has fees)
    total_no_fees = sum(projected_account_value_no_fees(a, assumptions) for a in accounts)
    total_fee_impact = total_no_fees - total_projected
    if total_fee_impact > 0:
        r_fee = total_row + 1
        _data_row(ws, r_fee, [
            "Lifetime cost of fees",
            "",
            "",
            total_fee_impact,
        ], bold=True, num_formats={4: GBP0})
        # Colour the fee impact value in a muted red
        ws.cell(row=r_fee, column=4).font = Font(name="Aptos", bold=True, color="DC2626", size=10)
        ws.cell(row=r_fee, column=1).font = Font(name="Aptos", bold=True, color="DC2626", size=10)

    # Assumptions block
    r = (total_row + 3) if total_fee_impact > 0 else (total_row + 2)
    ws.cell(row=r, column=1, value="Assumptions").font = _ACCENT_FONT
    for label, val in [
        ("Current age", int(current_age)),
        ("Retirement age", int(retirement_age)),
        ("Annual growth rate", f"{growth_rate*100:.1f}%"),
        ("Years to retirement", f"{exact_years:.1f}"),
    ]:
        r += 1
        ws.cell(row=r, column=1, value=label).font = _SUBTITLE_FONT
        ws.cell(row=r, column=2, value=val).font = _DATA_FONT

    # ── Sheet 2: Year by year ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Year by Year")
    _title_cell(ws2, 1, "Shelly Finance — Year-by-Year Projection", 3)
    _header_row(ws2, 3, ["Age", "Year", "Projected Total"])
    _set_col_width(ws2, 1, 10)
    _set_col_width(ws2, 2, 10)
    _set_col_width(ws2, 3, 22)

    curr_year = date.today().year
    for yr in range(0, whole_years + 1):
        age = int(current_age + yr)
        total = sum(
            projected_account_value_at_year(a, assumptions, yr)
            for a in accounts
        )
        yby_year_label = f"{curr_year + yr} (today)" if yr == 0 else curr_year + yr
        _data_row(ws2, yr + 4, [age, yby_year_label, total], num_formats={3: GBP0})

    # Final fractional-year point (matches summary card exactly)
    if exact_years > whole_years:
        final_row = whole_years + 4 + 1
        _data_row(ws2, final_row, [
            int(retirement_age),
            curr_year + whole_years + 1,
            total_projected,
        ], bold=True, num_formats={3: GBP0})

    # ── Sheet 3: Month by month (total portfolio) ────────────────────────────
    ws3 = wb.create_sheet("Month by Month")
    _title_cell(ws3, 1, "Shelly Finance — Monthly Projection", 3)
    _header_row(ws3, 3, ["Month", "Projected Total"])
    _set_col_width(ws3, 1, 16)
    _set_col_width(ws3, 2, 22)

    total_months = int(exact_years * 12)
    today = date.today()
    for m in range(0, total_months + 1):
        month_date = date(today.year + (today.month - 1 + m) // 12,
                          (today.month - 1 + m) % 12 + 1, 1)
        month_label = f"{month_date.strftime('%b %Y')}"
        if m == 0:
            month_label += " (today)"
        total = sum(
            projected_account_value_at_month(a, assumptions, m)
            for a in accounts
        )
        _data_row(ws3, m + 4, [month_label, total], num_formats={2: GBP0})

    # Final row at retirement
    _data_row(ws3, total_months + 5, ["Retirement", total_projected], bold=True, num_formats={2: GBP0})

    # ── Per-account sheets: year-by-year for each account ─────────────────
    for acc in accounts:
        # Sanitise name for Excel sheet title (max 31 chars, no special chars)
        safe_name = acc["name"][:28].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "(").replace("]", ")")
        ws_acc = wb.create_sheet(safe_name)

        acc_growth = account_growth_rate(acc, assumptions)
        acc_gross = account_gross_growth_rate(acc, assumptions)
        acc_fee_pct = effective_fee_pct(acc)
        acc_platform_pct = to_float(_safe_get(acc, "platform_fee_pct", 0))
        acc_platform_flat = to_float(_safe_get(acc, "platform_fee_flat", 0))
        acc_platform_cap = to_float(_safe_get(acc, "platform_fee_cap", 0))
        acc_fund_pct = to_float(_safe_get(acc, "fund_fee_pct", 0))
        acc_contribution_fee_pct = to_float(_safe_get(acc, "contribution_fee_pct", 0))
        acc_monthly = effective_monthly_contribution(acc, assumptions)
        acc_current = to_float(acc["current_value"])
        acc_projected = projected_account_value(acc, assumptions)
        acc_projected_no_fees = projected_account_value_no_fees(acc, assumptions)
        acc_fee_impact = acc_projected_no_fees - acc_projected
        acc_breakdown = contribution_breakdown(acc, assumptions)
        acc_contrib_fee_monthly = to_float(acc_breakdown.get("contribution_fee", 0))
        acc_total_months_for_fees = int(exact_years * 12)
        acc_total_contrib_fees = acc_contrib_fee_monthly * acc_total_months_for_fees
        has_annual_fees = acc_fee_pct > 0
        has_contrib_fee = acc_contribution_fee_pct > 0
        has_fees = has_annual_fees or has_contrib_fee

        max_cols = 7 if has_annual_fees else (6 if has_contrib_fee else 5)
        _title_cell(ws_acc, 1, f"Shelly Finance — {acc['name']}", max_cols)
        sub = ws_acc.cell(row=2, column=1, value=f"{acc['wrapper_type']} · {acc.get('provider') or ''}")
        sub.font = _SUBTITLE_FONT

        # Account summary
        _header_row(ws_acc, 4, ["", "Value"])
        _set_col_width(ws_acc, 1, 28)
        _set_col_width(ws_acc, 2, 18)
        summary_rows = [
            ("Current value", acc_current, GBP),
            ("You pay (monthly)", acc_breakdown["personal"], GBP),
        ]
        if has_contrib_fee:
            summary_rows.append(("Contribution fee deducted (monthly)", -acc_contrib_fee_monthly, GBP))
        summary_rows += [
            ("Total into pot (monthly)", acc_monthly, GBP),
            ("Growth rate (net of fees)", f"{acc_growth*100:.1f}%", None),
        ]
        if has_annual_fees:
            summary_rows.append(("Growth rate (gross)", f"{acc_gross*100:.1f}%", None))
            # Show granular fee breakdown if available
            if acc_platform_pct > 0:
                cap_note = f" (capped £{acc_platform_cap:,.0f}/yr)" if acc_platform_cap > 0 else ""
                summary_rows.append(("Platform fee", f"{acc_platform_pct:.2f}%{cap_note}", None))
            if acc_platform_flat > 0:
                summary_rows.append(("Platform fee (flat)", f"£{acc_platform_flat:,.0f}/yr", None))
            if acc_fund_pct > 0:
                summary_rows.append(("Fund fee (OCF)", f"{acc_fund_pct:.2f}%", None))
            summary_rows.append(("Total effective annual fee", f"{acc_fee_pct:.2f}%", None))
        if has_contrib_fee:
            summary_rows.append(("Contribution fee", f"{acc_contribution_fee_pct:.2f}% per contribution", None))
            summary_rows.append(("Total contribution fees paid", acc_total_contrib_fees, GBP0))
        summary_rows.append(("Projected at retirement", acc_projected, GBP0))
        if has_annual_fees:
            summary_rows.append(("Value without annual fees", acc_projected_no_fees, GBP0))
            summary_rows.append(("Lifetime cost of annual fees", acc_fee_impact, GBP0))

        for ri, (label, val, fmt) in enumerate(summary_rows, 5):
            _data_row(ws_acc, ri, [label, val], num_formats={2: fmt} if fmt else {})

        # Year-by-year table — columns vary by which fees apply
        yby_start = 5 + len(summary_rows) + 1
        if has_annual_fees and has_contrib_fee:
            yby_headers = ["Age", "Year", "Projected Value", "Growth", "Contributions", "Contrib. Fee (yr)", "Value (no ann. fees)", "Ann. Fee Impact"]
        elif has_annual_fees:
            yby_headers = ["Age", "Year", "Projected Value", "Growth", "Contributions", "Value (no fees)", "Fee Impact"]
        elif has_contrib_fee:
            yby_headers = ["Age", "Year", "Projected Value", "Growth", "Contributions", "Contrib. Fee (yr)"]
        else:
            yby_headers = ["Age", "Year", "Projected Value", "Growth", "Contributions"]
        _header_row(ws_acc, yby_start, yby_headers)
        _set_col_width(ws_acc, 3, 22)
        _set_col_width(ws_acc, 4, 18)
        _set_col_width(ws_acc, 5, 18)
        if len(yby_headers) >= 6:
            _set_col_width(ws_acc, 6, 22)
        if len(yby_headers) >= 7:
            _set_col_width(ws_acc, 7, 22)
        if len(yby_headers) >= 8:
            _set_col_width(ws_acc, 8, 18)

        is_lisa = acc.get("wrapper_type") == "Lifetime ISA"
        prev_val = acc_current
        for yr in range(0, whole_years + 1):
            age = int(current_age + yr)
            val = projected_account_value_at_year(acc, assumptions, yr)
            if yr == 0:
                contrib_this_year = 0
                contrib_fee_this_year = 0
            elif is_lisa and (current_age + yr) > 50:
                contrib_this_year = 0
                contrib_fee_this_year = 0
            else:
                contrib_this_year = acc_monthly * 12
                contrib_fee_this_year = acc_contrib_fee_monthly * 12
            growth_this_year = (val - prev_val - contrib_this_year) if yr > 0 else 0
            year_label = f"{curr_year + yr} (today)" if yr == 0 else curr_year + yr

            if has_annual_fees and has_contrib_fee:
                val_no_fees = projected_account_value_at_year_no_fees(acc, assumptions, yr)
                _data_row(ws_acc, yby_start + 1 + yr, [
                    age, year_label, val, growth_this_year, contrib_this_year,
                    contrib_fee_this_year, val_no_fees, val_no_fees - val,
                ], num_formats={3: GBP0, 4: GBP0, 5: GBP0, 6: GBP0, 7: GBP0, 8: GBP0})
            elif has_annual_fees:
                val_no_fees = projected_account_value_at_year_no_fees(acc, assumptions, yr)
                _data_row(ws_acc, yby_start + 1 + yr, [
                    age, year_label, val, growth_this_year, contrib_this_year,
                    val_no_fees, val_no_fees - val,
                ], num_formats={3: GBP0, 4: GBP0, 5: GBP0, 6: GBP0, 7: GBP0})
            elif has_contrib_fee:
                _data_row(ws_acc, yby_start + 1 + yr, [
                    age, year_label, val, growth_this_year, contrib_this_year, contrib_fee_this_year,
                ], num_formats={3: GBP0, 4: GBP0, 5: GBP0, 6: GBP0})
            else:
                _data_row(ws_acc, yby_start + 1 + yr, [
                    age, year_label, val, growth_this_year, contrib_this_year,
                ], num_formats={3: GBP0, 4: GBP0, 5: GBP0})
            prev_val = val

        # Final fractional-year row
        if exact_years > whole_years:
            final_r = yby_start + 1 + whole_years + 1
            if has_annual_fees and has_contrib_fee:
                _data_row(ws_acc, final_r, [
                    int(retirement_age), curr_year + whole_years + 1, acc_projected, "", "", "",
                    acc_projected_no_fees, acc_fee_impact,
                ], bold=True, num_formats={3: GBP0, 7: GBP0, 8: GBP0})
            elif has_annual_fees:
                _data_row(ws_acc, final_r, [
                    int(retirement_age), curr_year + whole_years + 1, acc_projected, "", "",
                    acc_projected_no_fees, acc_fee_impact,
                ], bold=True, num_formats={3: GBP0, 6: GBP0, 7: GBP0})
            elif has_contrib_fee:
                _data_row(ws_acc, final_r, [
                    int(retirement_age), curr_year + whole_years + 1, acc_projected, "", "", "",
                ], bold=True, num_formats={3: GBP0})
            else:
                _data_row(ws_acc, final_r, [
                    int(retirement_age), curr_year + whole_years + 1, acc_projected, "", "",
                ], bold=True, num_formats={3: GBP0})

        # ── Monthly breakdown table ──────────────────────────────────
        yearly_end = yby_start + 1 + whole_years + (2 if exact_years > whole_years else 1)
        mby_start = yearly_end + 2  # gap of 1 empty row

        if has_annual_fees and has_contrib_fee:
            _header_row(ws_acc, mby_start, ["Month", "Projected Value", "Contrib. Fee (mo)", "Value (no ann. fees)", "Ann. Fee Impact"])
        elif has_annual_fees:
            _header_row(ws_acc, mby_start, ["Month", "Projected Value", "Value (no fees)", "Fee Impact"])
        elif has_contrib_fee:
            _header_row(ws_acc, mby_start, ["Month", "Projected Value", "Contrib. Fee (mo)"])
        else:
            _header_row(ws_acc, mby_start, ["Month", "Projected Value"])

        acc_total_months = int(exact_years * 12)
        for m in range(0, acc_total_months + 1):
            m_date = date(today.year + (today.month - 1 + m) // 12,
                          (today.month - 1 + m) % 12 + 1, 1)
            m_label = f"{m_date.strftime('%b %Y')}"
            if m == 0:
                m_label += " (today)"
            m_val = projected_account_value_at_month(acc, assumptions, m)
            # Contribution fee is constant each month (except month 0)
            m_contrib_fee = acc_contrib_fee_monthly if m > 0 else 0.0

            if has_annual_fees and has_contrib_fee:
                m_val_nf = projected_account_value_at_month_no_fees(acc, assumptions, m)
                _data_row(ws_acc, mby_start + 1 + m, [
                    m_label, m_val, m_contrib_fee, m_val_nf, m_val_nf - m_val,
                ], num_formats={2: GBP0, 3: GBP0, 4: GBP0, 5: GBP0})
            elif has_annual_fees:
                m_val_nf = projected_account_value_at_month_no_fees(acc, assumptions, m)
                _data_row(ws_acc, mby_start + 1 + m, [
                    m_label, m_val, m_val_nf, m_val_nf - m_val,
                ], num_formats={2: GBP0, 3: GBP0, 4: GBP0})
            elif has_contrib_fee:
                _data_row(ws_acc, mby_start + 1 + m, [
                    m_label, m_val, m_contrib_fee,
                ], num_formats={2: GBP0, 3: GBP0})
            else:
                _data_row(ws_acc, mby_start + 1 + m, [
                    m_label, m_val,
                ], num_formats={2: GBP0})

        # Final retirement row
        m_final_r = mby_start + 1 + acc_total_months + 1
        if has_annual_fees and has_contrib_fee:
            _data_row(ws_acc, m_final_r, [
                "Retirement", acc_projected, acc_total_contrib_fees, acc_projected_no_fees, acc_fee_impact,
            ], bold=True, num_formats={2: GBP0, 3: GBP0, 4: GBP0, 5: GBP0})
        elif has_annual_fees:
            _data_row(ws_acc, m_final_r, [
                "Retirement", acc_projected, acc_projected_no_fees, acc_fee_impact,
            ], bold=True, num_formats={2: GBP0, 3: GBP0, 4: GBP0})
        elif has_contrib_fee:
            _data_row(ws_acc, m_final_r, [
                "Retirement", acc_projected, acc_total_contrib_fees,
            ], bold=True, num_formats={2: GBP0, 3: GBP0})
        else:
            _data_row(ws_acc, m_final_r, [
                "Retirement", acc_projected,
            ], bold=True, num_formats={2: GBP0})

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"projections_{date.today().isoformat()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Budget export ─────────────────────────────────────────────────────────────

GBP = '£#,##0.00'


def _income_key(db_sections):
    income_section = next((s for s in db_sections if "income" in s["key"].lower()), None)
    return income_section["key"] if income_section else (db_sections[0]["key"] if db_sections else "income")


def _write_budget_month_sheet(ws, title_text, db_sections, items, entry_map, item_id_col=False):
    """Render one month's budget into `ws`. Returns a dict of section_key → total.

    If item_id_col is True, an extra hidden column A carries the budget_item_id
    (used by the annual export so a future re-upload can match by ID).
    """
    col_offset = 1 if item_id_col else 0
    _set_col_width(ws, 1 + col_offset, 30)
    _set_col_width(ws, 2 + col_offset, 20)
    _set_col_width(ws, 3 + col_offset, 16)
    if item_id_col:
        _set_col_width(ws, 1, 8)

    _title_cell(ws, 1, title_text, 3 + col_offset)
    cell = ws.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%d %b %Y at %H:%M')}")
    cell.font = _SUBTITLE_FONT

    row = 4
    section_totals = {}

    for sec in db_sections:
        section_items = [it for it in items if it["section"] == sec["key"]]
        if not section_items:
            continue

        header_vals = ([""] if item_id_col else []) + [sec["label"], "", "Amount"]
        _header_row(ws, row, header_vals)
        row += 1

        sec_total = 0.0
        for item in section_items:
            amount = float(entry_map[item["id"]]["amount"]) if item["id"] in entry_map else float(item["default_amount"] or 0)
            vals = ([item["id"]] if item_id_col else []) + [item["name"], item["notes"] or "", amount]
            _data_row(ws, row, vals, num_formats={3 + col_offset: GBP})
            sec_total += amount
            row += 1

        total_vals = ([""] if item_id_col else []) + ["", "Section total", sec_total]
        _data_row(ws, row, total_vals, bold=True, num_formats={3 + col_offset: GBP})
        section_totals[sec["key"]] = sec_total
        row += 2

    total_income = section_totals.get(_income_key(db_sections), 0)
    total_expenses = sum(v for k, v in section_totals.items() if k != _income_key(db_sections))
    surplus = total_income - total_expenses

    for label, val in [("Total Income", total_income), ("Total Expenses", total_expenses), ("Surplus", surplus)]:
        row_vals = ([""] if item_id_col else []) + [label, "", val]
        _data_row(ws, row, row_vals, bold=(label == "Surplus"), num_formats={3 + col_offset: GBP})
        row += 1

    return section_totals


@export_bp.route("/budget/export.xlsx")
@login_required
def export_budget():
    uid = current_user.id
    month_key = valid_month_key(request.args.get("month")) or date.today().strftime("%Y-%m")
    month_label = datetime.strptime(month_key, "%Y-%m").strftime("%B %Y")

    db_sections = fetch_budget_sections(uid)
    items = fetch_budget_items(uid)
    entries = fetch_budget_entries(month_key, uid)
    entry_map = {e["budget_item_id"]: e for e in entries}
    if not entry_map:
        prior = fetch_prior_month_budget_entries(month_key, uid)
        entry_map = {e["budget_item_id"]: e for e in prior}

    wb = Workbook()
    ws = wb.active
    ws.title = f"Budget {month_key}"
    _write_budget_month_sheet(ws, f"Shelly Finance — Budget for {month_label}", db_sections, items, entry_map)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"budget_{month_key}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Annual budget export (UK tax year) ────────────────────────────────────────

_MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _tax_year_months(start_year):
    """Return the 12 month_key strings ('YYYY-MM') covering tax year start_year/start_year+1.

    UK tax year: Apr (start_year) through Mar (start_year + 1).
    """
    result = []
    for offset in range(12):
        m = 4 + offset
        y = start_year + (m - 1) // 12
        month = ((m - 1) % 12) + 1
        result.append(f"{y:04d}-{month:02d}")
    return result


def _resolved_month_map(month_key, uid, carry_forward):
    """Entry map for a month. If the month has no entries, fall back to the most
    recent prior month in-workbook (carry_forward), then to DB prior-month, then
    to default_amount (handled by the sheet writer)."""
    entries = fetch_budget_entries(month_key, uid)
    entry_map = {e["budget_item_id"]: e for e in entries}
    if not entry_map and carry_forward:
        # Reuse the last month we already resolved so the year stays consistent
        return dict(carry_forward)
    if not entry_map:
        prior = fetch_prior_month_budget_entries(month_key, uid)
        entry_map = {e["budget_item_id"]: e for e in prior}
    return entry_map


def _write_annual_summary_sheet(ws, months, month_labels, db_sections, items, month_entry_maps):
    """Write a wide 'items × 12 months + Total' matrix."""
    n_months = len(months)
    _set_col_width(ws, 1, 30)
    _set_col_width(ws, 2, 20)
    for col in range(3, 3 + n_months):
        _set_col_width(ws, col, 12)
    _set_col_width(ws, 3 + n_months, 14)

    _title_cell(ws, 1, "Shelly Finance — Annual Budget Summary", 3 + n_months)
    ws.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%d %b %Y at %H:%M')}").font = _SUBTITLE_FONT

    _header_row(ws, 4, ["Item", "Notes"] + month_labels + ["Total"])
    row = 5
    month_totals = [0.0] * n_months
    section_totals_per_month = {sec["key"]: [0.0] * n_months for sec in db_sections}

    for sec in db_sections:
        section_items = [it for it in items if it["section"] == sec["key"]]
        if not section_items:
            continue

        _header_row(ws, row, [sec["label"], ""] + [""] * n_months + [""])
        row += 1

        for item in section_items:
            row_amounts = []
            for i, month_key in enumerate(months):
                entry_map = month_entry_maps[month_key]
                amt = float(entry_map[item["id"]]["amount"]) if item["id"] in entry_map else float(item["default_amount"] or 0)
                row_amounts.append(amt)
                month_totals[i] += amt
                section_totals_per_month[sec["key"]][i] += amt
            item_total = sum(row_amounts)
            num_formats = {c: GBP for c in range(3, 4 + n_months)}
            _data_row(ws, row, [item["name"], item["notes"] or ""] + row_amounts + [item_total], num_formats=num_formats)
            row += 1

        sec_totals = section_totals_per_month[sec["key"]]
        num_formats = {c: GBP for c in range(3, 4 + n_months)}
        _data_row(ws, row, ["", "Section total"] + sec_totals + [sum(sec_totals)],
                  bold=True, num_formats=num_formats)
        row += 2

    # Overall summary
    income_key = _income_key(db_sections)
    income_by_month = section_totals_per_month.get(income_key, [0.0] * n_months)
    expense_by_month = [sum(section_totals_per_month[k][i] for k in section_totals_per_month if k != income_key)
                        for i in range(n_months)]
    surplus_by_month = [income_by_month[i] - expense_by_month[i] for i in range(n_months)]

    num_formats = {c: GBP for c in range(3, 4 + n_months)}
    for label, series in [("Total Income", income_by_month),
                          ("Total Expenses", expense_by_month),
                          ("Surplus", surplus_by_month)]:
        _data_row(ws, row, [label, ""] + series + [sum(series)],
                  bold=(label == "Surplus"), num_formats=num_formats)
        row += 1


def _write_investment_tracking_sheet(ws, uid, start_year, accounts, items, month_entry_maps, assumptions):
    """Planned (budget) vs Actual (logged) vs Allowance per wrapper, plus per-account detail."""
    _set_col_width(ws, 1, 26)
    _set_col_width(ws, 2, 16)
    _set_col_width(ws, 3, 16)
    _set_col_width(ws, 4, 16)
    _set_col_width(ws, 5, 16)
    _set_col_width(ws, 6, 12)

    ty_label = f"{start_year}/{str(start_year + 1)[-2:]}"
    _title_cell(ws, 1, f"Shelly Finance — Investment Tracking (Tax Year {ty_label})", 6)
    ws.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%d %b %Y at %H:%M')}").font = _SUBTITLE_FONT

    account_map = {int(a["id"]): dict(a) for a in accounts}

    # Planned per account = sum across 12 months of linked budget item amounts
    linked_items = [it for it in items if it.get("linked_account_id")]
    planned_per_account = {}
    for it in linked_items:
        aid = int(it["linked_account_id"])
        total = 0.0
        for mk, entry_map in month_entry_maps.items():
            total += float(entry_map[it["id"]]["amount"]) if it["id"] in entry_map else float(it["default_amount"] or 0)
        planned_per_account[aid] = planned_per_account.get(aid, 0.0) + total

    # Actual logged contributions in the tax year
    ty_start_iso = date(start_year, 4, 6).isoformat()
    ty_end_iso = date(start_year + 1, 4, 5).isoformat()
    isa_logs = fetch_isa_contributions(uid, ty_start_iso, ty_end_iso) or []
    pension_logs = fetch_pension_contributions(uid, ty_start_iso, ty_end_iso) or []

    actual_per_account = {}
    for row in list(isa_logs) + list(pension_logs):
        aid = int(row["account_id"])
        actual_per_account[aid] = actual_per_account.get(aid, 0.0) + float(row["amount"] or 0)

    # ── Wrapper rollup ───────────────────────────────────────────────────────
    isa_allowance = float(assumptions["isa_allowance"]) if assumptions and assumptions.get("isa_allowance") else 20000.0
    lisa_allowance = float(assumptions["lisa_allowance"]) if assumptions and assumptions.get("lisa_allowance") else 4000.0
    pension_allowance = float(assumptions["pension_allowance"]) if assumptions and assumptions.get("pension_allowance") else 60000.0

    def _sum(pred):
        planned = sum(planned_per_account.get(aid, 0.0) for aid in account_map if pred(account_map[aid]))
        actual = sum(actual_per_account.get(aid, 0.0) for aid in account_map if pred(account_map[aid]))
        return planned, actual

    isa_planned, isa_actual = _sum(lambda a: (a.get("wrapper_type") or "") in ISA_WRAPPER_TYPES)
    lisa_planned, lisa_actual = _sum(lambda a: (a.get("wrapper_type") or "") in LISA_WRAPPER_TYPES)
    pension_planned, pension_actual = _sum(lambda a: is_pension_account(a))

    _header_row(ws, 4, ["Wrapper", "Planned (budget)", "Logged (actuals)", "Allowance", "Remaining", "% Used"])

    def _pct(used, allowance):
        return (used / allowance * 100.0) if allowance > 0 else 0.0

    PCT = '0.0"%"'
    rows_data = [
        ("ISA (all)", isa_planned, isa_actual, isa_allowance),
        ("  of which LISA", lisa_planned, lisa_actual, lisa_allowance),
        ("Pension", pension_planned, pension_actual, pension_allowance),
    ]
    row = 5
    for label, planned, actual, allowance in rows_data:
        used = max(planned, actual)  # show worst-case usage for "remaining"
        remaining = max(allowance - used, 0.0)
        _data_row(ws, row, [label, planned, actual, allowance, remaining, _pct(used, allowance)],
                  num_formats={2: GBP, 3: GBP, 4: GBP, 5: GBP, 6: PCT})
        row += 1

    # ISA allowance note (LISA sits inside ISA £20k)
    note = ws.cell(row=row + 1, column=1,
                   value="Note: LISA contributions (£4k cap) count toward the overall ISA £20k allowance.")
    note.font = _SUBTITLE_FONT
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=6)
    row += 3

    # ── Per-account detail ────────────────────────────────────────────────────
    _header_row(ws, row, ["Account", "Wrapper", "Planned (annual)", "Logged (annual)", "Diff (Logged − Planned)", ""])
    row += 1

    for aid, acc in sorted(account_map.items(), key=lambda kv: kv[1].get("name") or ""):
        planned = planned_per_account.get(aid, 0.0)
        actual = actual_per_account.get(aid, 0.0)
        if planned == 0 and actual == 0:
            continue
        wrapper = acc.get("wrapper_type") or "—"
        _data_row(ws, row, [acc.get("name") or "", wrapper, planned, actual, actual - planned, ""],
                  num_formats={3: GBP, 4: GBP, 5: GBP})
        row += 1


@export_bp.route("/budget/annual-export.xlsx")
@login_required
def export_budget_annual():
    """Annual budget export: 12 month tabs (Apr→Mar of UK tax year) + Summary
    + Investment Tracking."""
    uid = current_user.id
    today = date.today()
    default_start = uk_tax_year_start(today).year
    try:
        start_year = int(request.args.get("tax_year_start") or default_start)
    except (ValueError, TypeError):
        start_year = default_start

    db_sections = fetch_budget_sections(uid)
    items = fetch_budget_items(uid)
    accounts = fetch_all_accounts(uid)
    assumptions = fetch_assumptions(uid)

    months = _tax_year_months(start_year)
    month_labels = []
    for mk in months:
        d = datetime.strptime(mk, "%Y-%m")
        month_labels.append(f"{_MONTH_NAMES[d.month - 1]} {d.year}")

    # Resolve per-month entry maps (with in-workbook carry-forward for empty months)
    month_entry_maps = {}
    carry = None
    for mk in months:
        em = _resolved_month_map(mk, uid, carry)
        month_entry_maps[mk] = em
        if em:
            carry = em

    wb = Workbook()
    # First sheet: Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _write_annual_summary_sheet(ws_sum, months, month_labels, db_sections, items, month_entry_maps)

    # 12 month sheets (re-uses the monthly format with hidden item_id column A)
    for mk, label in zip(months, month_labels):
        ws = wb.create_sheet(label)
        _write_budget_month_sheet(ws, f"Shelly Finance — Budget for {label}", db_sections, items,
                                  month_entry_maps[mk], item_id_col=True)

    # Investment Tracking
    ws_inv = wb.create_sheet("Investment Tracking")
    _write_investment_tracking_sheet(ws_inv, uid, start_year, accounts, items, month_entry_maps, assumptions)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"budget_tax_year_{start_year}-{str(start_year + 1)[-2:]}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Performance export ────────────────────────────────────────────────────────

@export_bp.route("/performance/export.xlsx")
@login_required
def export_performance():
    uid = current_user.id
    assumptions = fetch_assumptions(uid)
    accounts = fetch_all_accounts(uid)
    account_id = request.args.get("account_id")

    assumed_rate = to_float(assumptions["annual_growth_rate"]) if assumptions else 0.07
    assumed_monthly_total = sum(to_float(a["monthly_contribution"]) for a in accounts)

    per_account_data = fetch_monthly_performance_data_by_account(uid)
    account_map = {int(a["id"]): a for a in accounts}

    selected_account_id = None
    if account_id:
        try:
            selected_account_id = int(account_id)
        except Exception:
            selected_account_id = None
        if selected_account_id not in account_map:
            selected_account_id = None

    perf_portfolio = None
    if selected_account_id is None:
        monthly_data = fetch_monthly_performance_data(uid)
        perf_portfolio = compute_performance_series(monthly_data, assumed_rate, assumed_monthly_total)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _set_col_width(ws, 1, 26)
    _set_col_width(ws, 2, 12)
    _set_col_width(ws, 3, 12)
    _set_col_width(ws, 4, 14)
    _set_col_width(ws, 5, 16)
    _set_col_width(ws, 6, 16)
    _set_col_width(ws, 7, 16)
    _set_col_width(ws, 8, 16)
    _set_col_width(ws, 9, 16)
    _set_col_width(ws, 10, 18)

    _title_cell(ws, 1, "Shelly Finance — Performance Report", 10)
    cell = ws.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%d %b %Y at %H:%M')}")
    cell.font = _SUBTITLE_FONT

    _header_row(ws, 4, [
        "Entity",
        "Start",
        "End",
        "Months",
        "Total Return",
        "Annualised",
        "Contributed",
        "Market Gain",
        "Vs Plan",
        "Current Value",
    ])

    GBP = '£#,##0.00'
    PCT = '0.00"%"'
    row = 5

    def _append_summary(entity_name, perf):
        nonlocal row
        if not perf:
            _data_row(ws, row, [entity_name, "", "", 0, "", "", "", "", "", ""], bold=True)
            row += 1
            return
        labels = perf.get("labels") or []
        start_m = labels[0] if labels else ""
        end_m = labels[-1] if labels else ""
        _data_row(ws, row, [
            entity_name,
            start_m,
            end_m,
            int(perf.get("n_months") or 0),
            float(perf.get("total_return") or 0),
            float(perf.get("annualised_return") or 0) if perf.get("annualised_return") is not None else None,
            float(perf.get("total_contributed") or 0),
            float(perf.get("total_market_gain") or 0),
            float(perf.get("vs_plan") or 0),
            float(perf.get("current_value") or 0),
        ], bold=True, num_formats={5: PCT, 6: PCT, 7: GBP, 8: GBP, 9: GBP, 10: GBP})
        row += 1

    if selected_account_id is None:
        _append_summary("Portfolio", perf_portfolio)

        for aid, payload in per_account_data.items():
            acc = account_map.get(aid)
            if not acc:
                continue
            assumed_monthly = to_float(acc.get("monthly_contribution", 0))
            perf_acc = compute_performance_series(payload["rows"], assumed_rate, assumed_monthly)
            _append_summary(payload["account_name"], perf_acc)
    else:
        payload = per_account_data.get(selected_account_id, {"account_name": account_map[selected_account_id]["name"], "rows": []})
        acc = account_map[selected_account_id]
        assumed_monthly = to_float(acc.get("monthly_contribution", 0))
        perf_acc = compute_performance_series(payload["rows"], assumed_rate, assumed_monthly)
        _append_summary(payload["account_name"], perf_acc)

    def _safe_sheet_title(base, used):
        s = (base or "Sheet").strip()
        s = "".join("-" if ch in (":", "\\", "/", "?", "*", "[", "]") else ch for ch in s)
        s = s.strip().strip("'")[:31]
        if not s:
            s = "Sheet"
        if s not in used:
            used.add(s)
            return s
        i = 2
        while True:
            suffix = f" {i}"
            candidate = (s[:31 - len(suffix)] + suffix)[:31]
            if candidate not in used:
                used.add(candidate)
                return candidate
            i += 1

    used_titles = {ws.title}

    def _add_detail_sheet(title, perf):
        ws_d = wb.create_sheet(_safe_sheet_title(title, used_titles))
        _set_col_width(ws_d, 1, 10)
        _set_col_width(ws_d, 2, 16)
        _set_col_width(ws_d, 3, 16)
        _set_col_width(ws_d, 4, 18)
        _set_col_width(ws_d, 5, 16)
        _set_col_width(ws_d, 6, 12)

        _title_cell(ws_d, 1, f"Shelly Finance — {title}", 6)
        sub = ws_d.cell(row=2, column=1, value=f"Assumed growth: {assumed_rate*100:.1f}%")
        sub.font = _SUBTITLE_FONT

        if not perf or not perf.get("table_rows"):
            ws_d.cell(row=4, column=1, value="Not enough data yet (need at least two monthly snapshots).").font = _DATA_FONT
            return

        _header_row(ws_d, 4, ["Month", "Opening", "Contributions", "Market Gain / Loss", "Closing", "Return"])
        rows_chrono = list(reversed(perf["table_rows"]))
        for i, r in enumerate(rows_chrono, 5):
            _data_row(ws_d, i, [
                r["month_key"],
                float(r["opening"]),
                float(r["contribution"]),
                float(r["market_gain"]),
                float(r["closing"]),
                float(r["return_pct"]),
            ], num_formats={2: GBP, 3: GBP, 4: GBP, 5: GBP, 6: PCT})

    if selected_account_id is None:
        _add_detail_sheet("Portfolio (Monthly)", perf_portfolio)

        for aid, payload in per_account_data.items():
            acc = account_map.get(aid)
            if not acc:
                continue
            assumed_monthly = to_float(acc.get("monthly_contribution", 0))
            perf_acc = compute_performance_series(payload["rows"], assumed_rate, assumed_monthly)
            _add_detail_sheet(f"{payload['account_name']} (Monthly)", perf_acc)
    else:
        payload = per_account_data.get(selected_account_id, {"account_name": account_map[selected_account_id]["name"], "rows": []})
        acc = account_map[selected_account_id]
        assumed_monthly = to_float(acc.get("monthly_contribution", 0))
        perf_acc = compute_performance_series(payload["rows"], assumed_rate, assumed_monthly)
        _add_detail_sheet(f"{payload['account_name']} (Monthly)", perf_acc)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    if selected_account_id is None:
        fname = f"performance_{date.today().isoformat()}.xlsx"
    else:
        safe = "".join(ch for ch in (account_map[selected_account_id]["name"] or "account") if ch.isalnum() or ch in (" ", "-", "_")).strip().replace(" ", "_")
        fname = f"performance_{safe}_{date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
