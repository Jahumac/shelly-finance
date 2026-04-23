"""End-to-end tests for the annual budget import preview/confirm/cancel flow."""
from io import BytesIO

from openpyxl import load_workbook


def _seed(app, uid):
    """Seed a minimal budget: 1 income item, 1 linked-ISA item, 1 account."""
    from app.models import get_connection, create_budget_item, upsert_budget_entry
    with app.app_context():
        with get_connection() as conn:
            conn.execute(
                "INSERT INTO budget_sections (user_id, key, label, sort_order) VALUES (?, 'income', 'Income', 1)",
                (uid,),
            )
            conn.execute(
                "INSERT INTO budget_sections (user_id, key, label, sort_order) VALUES (?, 'inv', 'Investments', 2)",
                (uid,),
            )
            conn.execute(
                """INSERT INTO accounts (user_id, name, wrapper_type, current_value, monthly_contribution, is_active)
                   VALUES (?, 'My ISA', 'Stocks & Shares ISA', 5000, 333, 1)""",
                (uid,),
            )
            conn.commit()
            aid = conn.execute("SELECT id FROM accounts WHERE user_id=?", (uid,)).fetchone()["id"]

        create_budget_item(
            {"name": "Salary", "section": "income", "default_amount": 3500,
             "linked_account_id": None, "notes": "", "sort_order": 1}, uid)
        isa_id = create_budget_item(
            {"name": "ISA", "section": "inv", "default_amount": 333,
             "linked_account_id": aid, "notes": "", "sort_order": 1}, uid)
        upsert_budget_entry("2026-04", isa_id, 333, uid)
        return aid, isa_id


def _export_and_edit(auth_client, edits):
    """Download annual export, apply {"Sheet Name": {"ISA": 450, ...}} edits, return file bytes."""
    r = auth_client.get("/budget/annual-export.xlsx")
    wb = load_workbook(BytesIO(r.data))
    for sheet_name, row_edits in edits.items():
        ws = wb[sheet_name]
        for r_idx in range(4, ws.max_row + 1):
            name = ws.cell(r_idx, 2).value
            if isinstance(name, str) and name in row_edits:
                ws.cell(r_idx, 4).value = row_edits[name]
    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


def test_annual_import_preview_does_not_write_db(app, auth_client, make_user):
    from app.models import get_connection
    uid, _, _ = make_user()
    _seed(app, uid)

    buf = _export_and_edit(auth_client, {"Jun 2026": {"ISA": 450}})
    resp = auth_client.post("/budget/annual-import",
                            data={"file": (buf, "t.xlsx")},
                            content_type="multipart/form-data")
    assert resp.status_code == 200
    assert b"Preview" in resp.data
    assert b"June 2026" in resp.data
    assert b"450" in resp.data

    with app.app_context():
        with get_connection() as conn:
            # No June entry written yet; no overrides yet
            row = conn.execute(
                "SELECT COUNT(*) c FROM budget_entries be JOIN budget_items bi ON bi.id=be.budget_item_id "
                "WHERE bi.user_id=? AND be.month_key='2026-06'", (uid,)).fetchone()
            assert row["c"] == 0
            assert conn.execute("SELECT COUNT(*) c FROM contribution_overrides").fetchone()["c"] == 0


def test_annual_import_confirm_writes_db_and_syncs_overrides(app, auth_client, make_user):
    from app.models import get_connection
    uid, _, _ = make_user()
    _seed(app, uid)

    buf = _export_and_edit(auth_client, {"Jun 2026": {"ISA": 450}, "Oct 2026": {"Salary": 3700}})
    auth_client.post("/budget/annual-import",
                     data={"file": (buf, "t.xlsx")},
                     content_type="multipart/form-data")
    resp = auth_client.post("/budget/annual-import/confirm")
    assert resp.status_code == 302

    with app.app_context():
        with get_connection() as conn:
            jun = conn.execute(
                "SELECT amount FROM budget_entries be JOIN budget_items bi ON bi.id=be.budget_item_id "
                "WHERE bi.user_id=? AND be.month_key='2026-06' AND bi.name='ISA'", (uid,)).fetchone()
            oct_sal = conn.execute(
                "SELECT amount FROM budget_entries be JOIN budget_items bi ON bi.id=be.budget_item_id "
                "WHERE bi.user_id=? AND be.month_key='2026-10' AND bi.name='Salary'", (uid,)).fetchone()
            override = conn.execute(
                "SELECT override_amount FROM contribution_overrides WHERE from_month='2026-06'").fetchone()
            assert jun["amount"] == 450
            assert oct_sal["amount"] == 3700
            assert override["override_amount"] == 450  # linked item → override synced


def test_annual_import_empty_diff_redirects_without_preview(app, auth_client, make_user):
    uid, _, _ = make_user()
    _seed(app, uid)

    buf = _export_and_edit(auth_client, {})  # no edits → matches current state
    resp = auth_client.post("/budget/annual-import",
                            data={"file": (buf, "t.xlsx")},
                            content_type="multipart/form-data",
                            follow_redirects=False)
    assert resp.status_code == 302
    assert "/budget/" in resp.headers["Location"]


def test_annual_import_cancel_discards_staged_changes(app, auth_client, make_user):
    from app.models import get_connection
    uid, _, _ = make_user()
    _seed(app, uid)

    buf = _export_and_edit(auth_client, {"Jun 2026": {"ISA": 450}})
    auth_client.post("/budget/annual-import",
                     data={"file": (buf, "t.xlsx")},
                     content_type="multipart/form-data")
    resp = auth_client.post("/budget/annual-import/cancel", follow_redirects=False)
    assert resp.status_code == 302

    with app.app_context():
        with get_connection() as conn:
            row = conn.execute(
                "SELECT COUNT(*) c FROM budget_entries be JOIN budget_items bi ON bi.id=be.budget_item_id "
                "WHERE bi.user_id=? AND be.month_key='2026-06'", (uid,)).fetchone()
            assert row["c"] == 0


def test_annual_import_stale_confirm_redirects_gracefully(app, auth_client, make_user):
    uid, _, _ = make_user()
    _seed(app, uid)

    resp = auth_client.post("/budget/annual-import/confirm", follow_redirects=False)
    assert resp.status_code == 302
